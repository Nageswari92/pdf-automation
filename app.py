import io
import os
import re
import openpyxl
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
import streamlit as st

# ============================================
# AUTOMATION CORE LOGIC
# ============================================

def process_bom(pdf_file, template_file):
    """
    Reads uploaded BOM PDF and populates the uploaded VSE Excel Template.
    Fills data up to row 1399 MAX without altering/deleting data below row 1399.
    Returns the processed Excel file as bytes.
    """
    # 1. Read PDF Text
    pdf_text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pdf_text += text + "\n"

    # 2. Extract Header Information
    header = {
        "Assembly Number": "",
        "Assembly Rev": "",
        "Description": "",
        "VSE Revision": "",
        "Customer ID": "",
        "Ship Site": ""
    }

    for line in pdf_text.split("\n"):
        line = line.strip()
        if line.startswith("Assembly Number"):
            header["Assembly Number"] = line.replace("Assembly Number", "").strip()
        elif line.startswith("Assembly Rev"):
            header["Assembly Rev"] = line.replace("Assembly Rev", "").strip()
        elif line.startswith("Description"):
            header["Description"] = line.replace("Description", "").strip()
        elif line.startswith("VSE Revision"):
            header["VSE Revision"] = line.replace("VSE Revision", "").strip()
        elif line.startswith("Customer ID"):
            header["Customer ID"] = line.replace("Customer ID", "").strip()
        elif line.startswith("Ship Site"):
            header["Ship Site"] = line.replace("Ship Site", "").strip()

    # 3. Extract Structured BOM Table
    table = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            extracted_table = page.extract_table()
            if extracted_table:
                table.extend(extracted_table)

    # 4. Clean BOM Table
    clean_bom = []
    for row in table:
        if not row:
            continue
        row = [cell if cell is not None else "" for cell in row]
        if row[0] in ["Lvl", ""]:
            continue
        clean_bom.append(row)

    # 5. Populate VSE Excel Template
    workbook = load_workbook(template_file)
    worksheet = workbook.active

    # Write Header Values
    worksheet["D2"] = header["Assembly Number"]
    worksheet["D3"] = header["Assembly Rev"]
    worksheet["D4"] = header["Description"]
    worksheet["D5"] = header["VSE Revision"]
    worksheet["D6"] = header["Customer ID"]
    worksheet["D7"] = header["Ship Site"]

    # Find Start Row for BOM
    bom_header_row = None
    for row in worksheet.iter_rows():
        if row[0].value == "Lvl":
            bom_header_row = row[0].row
            break

    start_row = (bom_header_row + 1) if bom_header_row else 10
    MAX_ROW_LIMIT = 1399  # Safety stop row to protect data from row 1400+

    # Write BOM Data safely up to row 1399
    for i, row in enumerate(clean_bom):
        excel_row = start_row + i

        # Prevent writing into row 1400 or beyond
        if excel_row > MAX_ROW_LIMIT:
            st.warning(f"⚠️ BOM contains more rows than space allows. Writing was stopped at Row {MAX_ROW_LIMIT} to preserve template formulas.")
            break

        for col in range(11):
            value = row[col] if col < len(row) else ""
            worksheet.cell(row=excel_row, column=col + 1).value = value

    # Save to In-Memory Buffer for Streamlit Download
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    return output_stream, clean_bom, header


def get_output_filename(uploaded_filename):
    """
    Generates output filename based on PDF filename:
    - Removes .pdf extension
    - Removes 'DC' (including '_DC' or '-DC')
    - Appends .xlsx extension
    """
    base_name = os.path.splitext(uploaded_filename)[0]
    base_name = re.sub(r'[-_]?DC', '', base_name, flags=re.IGNORECASE).strip('_- ')
    return f"{base_name}.xlsx"


# ============================================
# STREAMLIT UI INTERFACE
# ============================================

st.set_page_config(page_title="BOM Automation Tool", layout="wide", page_icon="⚙️")

st.title("⚙️ BOM TO VSE AUTOMATION TOOL")
st.write("Upload your BOM PDF and VSE Template Excel file to generate the processed output.")

st.divider()

# File Uploaders
col1, col2 = st.columns(2)

with col1:
    uploaded_pdf = st.file_uploader("📄 Upload BOM PDF File", type=["pdf"])

with col2:
    uploaded_template = st.file_uploader("📊 Upload VSE Template File", type=["xlsx"])

st.divider()

if st.button("🚀 RUN AUTOMATION", type="primary", use_container_width=True):
    if uploaded_pdf is None:
        st.error("Please upload a BOM PDF File!")
    elif uploaded_template is None:
        st.error("Please upload a VSE Template File!")
    else:
        with st.spinner("Processing BOM Data & Populating Template..."):
            try:
                # Dynamic Output Filename
                output_filename = get_output_filename(uploaded_pdf.name)

                # Process
                excel_bytes, clean_bom, header = process_bom(uploaded_pdf, uploaded_template)

                st.success("✅ VSE File Generated Successfully!")
                st.info(f"📁 Generated File Name: **{output_filename}**")

                # Header Details
                st.subheader("📋 Extracted Header Details")
                st.json(header)

                # Data Preview
                st.subheader("🔍 BOM Data Preview")
                if clean_bom:
                    cols = ["Lvl", "Item #", "VSE P/N", "Qty", "Customer P/N", "Rev", "Description", "Ref Des", "UOM", "Mfr", "Mfr P/N"]
                    df_preview = pd.DataFrame(clean_bom, columns=cols[:len(clean_bom[0])])
                    st.dataframe(df_preview, use_container_width=True)

                st.divider()

                # Download Button with Dynamic Filename
                st.download_button(
                    label=f"📥 DOWNLOAD {output_filename}",
                    data=excel_bytes,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            except Exception as e:
                st.error(f"❌ Error occurred during processing: {e}")
